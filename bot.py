#!/usr/bin/env python3
"""
wsotp_automator_status_excel.py ✅ (Choreo Ready Version)
"""
import os
import asyncio
import json
from telethon import TelegramClient, events, errors
from openpyxl import Workbook, load_workbook

# ========== CONFIG ==========
API_ID = 21240549
API_HASH = "82d24135c6212258f861d76ea39bab7a"
BOT_TOKEN = "8046319646:AAHyQNHs56OKA09Wwx0ebD2Gi77nnAkrID0"
TARGET_BOT = "@wsotp200bot"

EXCEL_FILE = "yellow_numbers.xlsx"
PERSIST_FILE = "yellow_saved.json"
USER_SESSION = "user_wsotp.session"
BOT_SESSION = "bot_wsotp.session"
DELAY_BETWEEN_SENDS = 0.5
# ============================

def reset_all():
    if os.path.exists(EXCEL_FILE):
        os.remove(EXCEL_FILE)
    if os.path.exists(PERSIST_FILE):
        os.remove(PERSIST_FILE)
    wb = Workbook()
    ws = wb.active
    ws.title = "yellow"
    ws.append(["Number", "Status"])
    wb.save(EXCEL_FILE)
    print("🧹 Reset complete: fresh Excel and JSON created.")
    return set()

def save_saved_set(data):
    with open(PERSIST_FILE, "w", encoding="utf-8") as f:
        json.dump(sorted(list(data)), f, ensure_ascii=False, indent=2)

def append_to_excel(number):
    try:
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        ws.append([number, "Not Registered ✅"])
        wb.save(EXCEL_FILE)
        print(f"[+] Saved {number} → Not Registered ✅")
    except Exception as e:
        print("Excel save error:", e)

user_client = TelegramClient(USER_SESSION, API_ID, API_HASH)
bot_client = TelegramClient(BOT_SESSION, API_ID, API_HASH)
yellow_set = set()

async def send_number_to_target(number):
    try:
        await user_client.send_message(TARGET_BOT, number)
        print(f"→ Sent: {number}")
    except errors.FloodWaitError as e:
        print(f"⏳ FloodWait {e.seconds}s... sleeping")
        await asyncio.sleep(e.seconds + 1)
        await user_client.send_message(TARGET_BOT, number)
    except Exception as e:
        print(f"Send error: {e}")

@bot_client.on(events.NewMessage)
async def bot_receive(event):
    global yellow_set
    try:
        if not event.is_private:
            return
        txt = (event.message.message or "").strip()
        if not txt:
            return
        if txt.startswith("/start"):
            await event.reply("✅ Automation চালু আছে। নাম্বার পাঠাও, পুরনো ডেটা রিসেট হবে।")
            return
        if txt.startswith("/export"):
            if os.path.exists(EXCEL_FILE):
                await bot_client.send_file(event.chat_id, EXCEL_FILE, caption="📤 Excel ফাইল")
            else:
                await event.reply("⚠️ কোনো Excel ফাইল তৈরি হয়নি।")
            return
        if txt.startswith("/status"):
            total = len(yellow_set)
            await event.reply(f"📊 মোট Not Registered পাওয়া গেছে: {total}")
            return
        parts = [p.strip() for p in txt.replace(",", " ").split() if p.strip()]
        if not parts:
            await event.reply("⚠️ কোনো নাম্বার পাওয়া যায়নি।")
            return
        try:
            await event.delete()
        except:
            pass
        yellow_set = reset_all()
        await bot_client.send_message(event.chat_id, "✅ Send successfully. Wait for results.")
        for num in parts:
            await send_number_to_target(num)
            await asyncio.sleep(DELAY_BETWEEN_SENDS)
    except Exception as e:
        print(f"⚠️ bot_receive error: {e}")

def extract_number(text):
    parts = text.split()
    for p in parts:
        if any(ch.isdigit() for ch in p):
            return p
    return text.strip()

async def process_not_registered(text):
    global yellow_set
    if "🟡" not in text and "Not Registered" not in text:
        return
    number = extract_number(text)
    if number in yellow_set:
        return
    yellow_set.add(number)
    save_saved_set(yellow_set)
    append_to_excel(number)

@user_client.on(events.NewMessage(chats=TARGET_BOT))
async def on_new_msg(event):
    await process_not_registered(event.message.message or "")

@user_client.on(events.MessageEdited(chats=TARGET_BOT))
async def on_edited(event):
    await process_not_registered(event.message.message or "")

async def main():
    print("🚀 Starting...")
    await user_client.start()
    me = await user_client.get_me()
    print(f"[user_client] Logged in as: {me.username or me.first_name}")
    await bot_client.start(bot_token=BOT_TOKEN)
    bot_me = await bot_client.get_me()
    print(f"[bot_client] Running as: @{bot_me.username}")
    print("✅ Ready! Send numbers to your bot. Old Excel auto-clears.")
    await asyncio.Future()

if __name__ == "__main__":
    try:
        asyncio.run(main())
    except KeyboardInterrupt:
        print("🛑 Saving...")
        save_saved_set(yellow_set)
